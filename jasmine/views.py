from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, Count
from django.utils import timezone
from datetime import timedelta
import calendar
from orders.models import Order
from userauths.models import User
from products.models import Product, OrderItem

def is_owner(user):
    return user.is_authenticated and user.is_superuser

@user_passes_test(is_owner, login_url='/auth/')
def dashboard_view(request):
    today = timezone.now()
    first_day_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # KPI 1: Total Revenue (Bulan Ini) - only completed or shipped
    revenue_this_month = Order.objects.filter(
        status__in=['shipped', 'completed'],
        created_at__gte=first_day_of_month
    ).aggregate(total=Sum('total'))['total'] or 0
    
    # KPI 2: Total Orders (Bulan Ini)
    orders_this_month = Order.objects.filter(
        created_at__gte=first_day_of_month
    ).count()
    
    # KPI 3: Customer Baru
    new_customers = User.objects.filter(
        date_joined__gte=first_day_of_month,
        is_staff=False,
        is_superuser=False
    ).count()
    
    # KPI 4: Top Products
    top_products = OrderItem.objects.filter(
        order__status__in=['paid', 'shipped', 'completed'],
        order__created_at__gte=first_day_of_month
    ).values('product__name').annotate(total_qty=Sum('quantity')).order_by('-total_qty')[:10]
    
    context = {
        'revenue_this_month': revenue_this_month,
        'orders_this_month': orders_this_month,
        'new_customers': new_customers,
        'top_products': top_products,
    }
    return render(request, 'jasmine/dashboard.html', context)

@user_passes_test(is_owner, login_url='/auth/')
def export_excel_view(request):
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill
    
    # Get filters if any, for now all
    orders = Order.objects.all().order_by('-created_at')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Transaksi"
    
    # Header
    headers = ["Nomor Pesanan", "Tanggal", "Pelanggan", "Total (Rp)", "Status", "Ekspedisi", "Resi"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
        
    for row_num, order in enumerate(orders, 2):
        ws.cell(row=row_num, column=1, value=order.order_number)
        ws.cell(row=row_num, column=2, value=order.created_at.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row_num, column=3, value=order.user.email if order.user else "Guest")
        ws.cell(row=row_num, column=4, value=float(order.total))
        ws.cell(row=row_num, column=5, value=order.get_status_display())
        ws.cell(row=row_num, column=6, value=f"{order.courier.upper()} {order.shipping_service}")
        ws.cell(row=row_num, column=7, value=order.tracking_number or "-")
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Laporan_ZTP_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response
