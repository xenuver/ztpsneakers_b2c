from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.utils import timezone
from .models import Order
import openpyxl
from openpyxl.styles import Font, PatternFill

@staff_member_required
def export_excel_admin_view(request):
    if not request.user.is_superuser:
        return HttpResponse("Akses ditolak. Fitur ini khusus Owner/Superuser.", status=403)
    
    orders = Order.objects.all().order_by('-created_at')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Transaksi"
    
    headers = ["Nomor Pesanan", "Tanggal", "Pelanggan", "Total (Rp)", "Status", "Ekspedisi", "Resi"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
        
    for row_num, order in enumerate(orders, 2):
        ws.cell(row=row_num, column=1, value=order.order_number)
        ws.cell(row=row_num, column=2, value=order.created_at.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row_num, column=3, value=order.user.email if order.user else "Guest")
        ws.cell(row=row_num, column=4, value=float(order.total))
        ws.cell(row=row_num, column=5, value=order.get_status_display())
        ws.cell(row=row_num, column=6, value=f"{order.courier.upper()} {order.shipping_service}")
        ws.cell(row=row_num, column=7, value=order.tracking_number or "-")
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Laporan_ZTP_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response
